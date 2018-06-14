#encoding=utf-8
from openpyxl import load_workbook

class ParseExcel(object):

    def __init__(self,excelPath,sheetName):
        self.wb=load_workbook(excelPath)
        self.sheet=self.wb.get_sheet_by_name(sheetName)
        self.maxRowNum=self.sheet.max_row

    def getDataFromSheet(self):
        dataList=[]
        for line in self.sheet.rows[1:]:
            tmpList=[]
            tmpList.append(line[1].value)
            tmpList.append(line[2].value)
            dataList.append(tmpList)
        return dataList

if __name__ == '__main__':
    excelPath=r"B:\test\TestData.xlsx"
    sheetName=u"test"
    pe=ParseExcel(excelPath,sheetName)

    for i in pe.getDataFromSheet():
        print i[0],i[1]